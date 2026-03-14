"""
modules/warehouse/routes.py — Склад: залишки, прихід, інвентаризація
Author: White
"""
import json
import os
import uuid
from datetime import datetime
from pathlib import Path
from flask import (
    Blueprint, render_template, request, redirect,
    url_for, jsonify, session, make_response, current_app
)
from core.auth import login_required
from core.db import get_connection
from core.audit import log_action
from core.hooks import emit
from core.warehouse import get_stock as _get_stock

ALLOWED_SCAN_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".tif"}

def _save_scan_file(file, doc_number: str = "", doc_date: str = "") -> tuple[str, str] | tuple[None, None]:
    """Зберегти завантажений файл скану приходу. Повертає (path, original_name) або (None, None)."""
    if not file or not file.filename:
        return None, None
    original_name = file.filename
    ext = Path(original_name).suffix.lower()
    if ext not in ALLOWED_SCAN_EXTENSIONS:
        return None, None
    from core.settings import get_storage_path
    scans_dir = get_storage_path() / "scans" / "income"
    scans_dir.mkdir(parents=True, exist_ok=True)
    # Осмислене ім'я: income_<номер>_<дата>.<ext> або з uuid якщо немає даних
    num_safe  = (doc_number or "").replace("/", "-").replace(" ", "_")
    date_safe = (doc_date  or "")[:10].replace("-", "")
    if num_safe or date_safe:
        parts = [p for p in ["income", num_safe, date_safe] if p]
        filename = "_".join(parts) + ext
    else:
        filename = f"income_{uuid.uuid4().hex}{ext}"
    save_path = scans_dir / filename
    file.save(str(save_path))
    return str(save_path), original_name

bp = Blueprint("warehouse", __name__, url_prefix="/warehouse")

# Категорії (римські цифри)
CATEGORIES = ["I", "II", "III"]


# ─────────────────────────────────────────────────────────────
#  Допоміжні функції
# ─────────────────────────────────────────────────────────────

def _get_items_with_serial(conn) -> list:
    """Словник майна — тільки позиції."""
    return conn.execute(
        "SELECT id, name, unit_of_measure, has_serial_number, is_inventory "
        "FROM item_dictionary ORDER BY name"
    ).fetchall()


def _get_doc_types(conn) -> list:
    return conn.execute(
        "SELECT id, name, short_name FROM document_types ORDER BY id"
    ).fetchall()



# ─────────────────────────────────────────────────────────────
#  Залишки (головна сторінка складу)
# ─────────────────────────────────────────────────────────────

@bp.route("/")
@login_required
def index():
    conn = get_connection()
    stock = _get_stock(conn)
    conn.close()

    # Загальна сума
    total_sum = sum(r["total_sum"] for r in stock)

    return render_template(
        "warehouse/index.html",
        stock=stock,
        total_sum=total_sum,
        categories=CATEGORIES,
    )


# ─────────────────────────────────────────────────────────────
#  ПРИХІД — список
# ─────────────────────────────────────────────────────────────

@bp.route("/income/")
@login_required
def income_list():
    conn = get_connection()

    # Підтверджені приходи з warehouse_income (групуємо по документу)
    confirmed = conn.execute("""
        SELECT MIN(wi.id) AS id, wi.date, wi.document_number, wi.supplier,
               wi.source_type, wi.notes, wi.scan_path, wi.scan_original_name,
               dt.name AS doc_type_name, wi.document_type_id,
               SUM(wi.quantity * wi.price) AS total_sum,
               COUNT(*) AS items_count,
               'confirmed' AS status,
               NULL AS doc_id,
               wi.created_at
        FROM warehouse_income wi
        LEFT JOIN document_types dt ON wi.document_type_id = dt.id
        WHERE COALESCE(wi.status,'confirmed') = 'confirmed'
        GROUP BY wi.document_number, wi.date, wi.supplier, wi.source_type,
                 wi.document_type_id, wi.scan_path, wi.created_at
        ORDER BY wi.date DESC, MIN(wi.id) DESC
    """).fetchall()

    # Чернетки з income_docs
    drafts = conn.execute("""
        SELECT d.id, d.date, d.document_number, d.supplier,
               d.source_type, d.notes, d.scan_path, d.scan_original_name,
               dt.name AS doc_type_name,
               COALESCE(SUM(i.quantity * i.price), 0) AS total_sum,
               COUNT(i.id) AS items_count,
               'draft' AS status,
               d.id AS doc_id,
               d.created_at
        FROM income_docs d
        LEFT JOIN document_types dt ON d.document_type_id = dt.id
        LEFT JOIN income_doc_items i ON i.doc_id = d.id
        WHERE d.status = 'draft'
        GROUP BY d.id
        ORDER BY d.created_at DESC
    """).fetchall()

    conn.close()

    # Об'єднуємо: чернетки зверху, потім підтверджені
    all_rows = [dict(r) for r in drafts] + [dict(r) for r in confirmed]
    return render_template("warehouse/income_list.html", rows=all_rows)


# ─────────────────────────────────────────────────────────────
#  ПРИХІД — форма
# ─────────────────────────────────────────────────────────────

def _get_norm_groups(conn) -> list:
    """Групи словника норм для вибору при додаванні позиції майна."""
    nd_rows = conn.execute("""
        SELECT g.id AS group_id, g.name AS group_name, g.sort_order,
               nd.id, nd.name, nd.unit, nd.sort_order AS item_order
        FROM norm_dict_groups g
        JOIN norm_dictionary nd ON nd.group_id = g.id
        WHERE g.is_active=1 AND nd.is_active=1
        ORDER BY g.sort_order, nd.sort_order
    """).fetchall()
    groups_dict = {}
    for r in nd_rows:
        gid = r["group_id"]
        if gid not in groups_dict:
            groups_dict[gid] = {"id": gid, "name": r["group_name"], "norms": []}
        groups_dict[gid]["norms"].append({
            "id": r["id"], "name": r["name"], "unit": r["unit"] or "шт",
        })
    return list(groups_dict.values())


def _parse_income_rows(form, save_as_draft: bool, conn) -> tuple[list, list]:
    """Розбирає рядки позицій з форми. Повертає (rows, errors)."""
    item_ids    = form.getlist("item_id[]")
    quantities  = form.getlist("quantity[]")
    prices      = form.getlist("price[]")
    cats        = form.getlist("category[]")
    nom_codes   = form.getlist("nom_code[]")
    serials_raw = form.getlist("serial_numbers[]")

    rows, errors = [], []
    for i, item_id in enumerate(item_ids):
        if not item_id:
            continue
        try:
            qty   = float(quantities[i]) if i < len(quantities) and quantities[i] else 0
            price = float(prices[i])     if i < len(prices)     and prices[i]     else 0
        except ValueError:
            if not save_as_draft:
                errors.append(f"Рядок {i+1}: невірна кількість або ціна")
            qty, price = 0, 0

        if qty <= 0 and not save_as_draft:
            errors.append(f"Рядок {i+1}: кількість має бути більше 0")
            continue

        cat      = cats[i]      if i < len(cats)      else "I"
        nom_code = nom_codes[i] if i < len(nom_codes) else ""
        s_raw    = serials_raw[i] if i < len(serials_raw) else ""

        serials = []
        if not save_as_draft:
            item_row = conn.execute(
                "SELECT has_serial_number FROM item_dictionary WHERE id=?", (item_id,)
            ).fetchone()
            if item_row and item_row["has_serial_number"]:
                serials = [s.strip() for s in s_raw.split(",") if s.strip()]
                if not serials:
                    errors.append(f"Рядок {i+1}: потрібно вказати серійний номер")
                    continue
                if len(serials) != int(qty):
                    errors.append(f"Рядок {i+1}: серійних номерів {len(serials)}, а позицій {int(qty)}")
                    continue

        rows.append({
            "item_id": item_id, "quantity": qty, "price": price,
            "category": cat, "nom_code": nom_code, "serials": serials,
        })
    return rows, errors


@bp.route("/income/new", methods=["GET", "POST"])
@login_required
def income_new():
    conn = get_connection()
    items       = _get_items_with_serial(conn)
    doc_types   = _get_doc_types(conn)
    norm_groups = _get_norm_groups(conn)

    if request.method == "POST":
        save_as_draft = request.form.get("save_as_draft") == "1"
        date        = request.form.get("date", "").strip()
        doc_number  = request.form.get("doc_number", "").strip()
        doc_type_id = request.form.get("doc_type_id") or None
        supplier    = request.form.get("supplier", "").strip()
        notes       = request.form.get("notes", "").strip()
        source_type = request.form.get("source_type", "income_doc")

        errors = []
        if not date and not save_as_draft:
            errors.append("Вкажіть дату")

        rows, row_errors = _parse_income_rows(request.form, save_as_draft, conn)
        errors.extend(row_errors)

        if not rows and not save_as_draft:
            errors.append("Додайте хоча б одну позицію")

        if errors:
            conn.close()
            return render_template("warehouse/income_form.html",
                items=items, doc_types=doc_types, categories=CATEGORIES,
                norm_groups=norm_groups, errors=errors, form=request.form)

        scan_path, scan_original_name = _save_scan_file(
            request.files.get("scan_file"), doc_number, date
        )

        if save_as_draft:
            # Зберігаємо в income_docs
            cur = conn.execute("""
                INSERT INTO income_docs
                    (status, date, document_number, document_type_id, supplier,
                     source_type, notes, scan_path, scan_original_name, created_by)
                VALUES ('draft',?,?,?,?,?,?,?,?,?)
            """, (date, doc_number, doc_type_id, supplier,
                  source_type, notes, scan_path, scan_original_name,
                  session.get("user_id")))
            doc_id = cur.lastrowid
            for row in rows:
                conn.execute("""
                    INSERT INTO income_doc_items
                        (doc_id, item_id, quantity, price, category, nom_code, serial_numbers)
                    VALUES (?,?,?,?,?,?,?)
                """, (doc_id, row["item_id"], row["quantity"], row["price"],
                      row["category"], row["nom_code"],
                      ",".join(row["serials"]) if row["serials"] else None))
            conn.commit()
            conn.close()
            from flask import flash
            flash("Чернетку збережено. Відкрийте її для редагування або підтвердження.", "info")
            return redirect(url_for("warehouse.income_list"))
        else:
            # Зберігаємо одразу в warehouse_income
            income_id = None
            for row in rows:
                cur = conn.execute("""
                    INSERT INTO warehouse_income
                        (date, document_number, document_type_id, supplier,
                         item_id, quantity, price, category, nom_code,
                         notes, source_type, status,
                         scan_path, scan_original_name, created_by, created_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,'confirmed',?,?,?,datetime('now','localtime'))
                """, (date, doc_number, doc_type_id, supplier,
                      row["item_id"], row["quantity"], row["price"],
                      row["category"], row["nom_code"], notes, source_type,
                      scan_path, scan_original_name, session.get("user_id")))
                income_id = cur.lastrowid
                for sn in row["serials"]:
                    conn.execute("""
                        INSERT INTO item_serials (item_id, serial_number, warehouse_income_id, status)
                        VALUES (?,?,?,'stock')
                    """, (row["item_id"], sn, income_id))
            conn.commit()
            log_action("add", "warehouse_income", income_id,
                       new_data={"date": date, "doc_number": doc_number, "rows": len(rows)})
            emit("warehouse.income", date=date, rows=rows)
            conn.close()
            return redirect(url_for("warehouse.income_list"))

    conn.close()
    today = datetime.now().strftime("%Y-%m-%d")
    return render_template("warehouse/income_form.html",
        items=items, doc_types=doc_types, categories=CATEGORIES,
        norm_groups=norm_groups, errors=[], form={"date": today})


@bp.route("/income/draft/<int:doc_id>/edit", methods=["GET", "POST"])
@login_required
def income_draft_edit(doc_id):
    """Редагування чернетки приходу."""
    conn = get_connection()
    doc = conn.execute("SELECT * FROM income_docs WHERE id=? AND status='draft'", (doc_id,)).fetchone()
    if not doc:
        conn.close()
        from flask import flash
        flash("Чернетку не знайдено", "danger")
        return redirect(url_for("warehouse.income_list"))

    items       = _get_items_with_serial(conn)
    doc_types   = _get_doc_types(conn)
    norm_groups = _get_norm_groups(conn)

    if request.method == "POST":
        save_as_draft = request.form.get("save_as_draft") == "1"
        confirm_now   = request.form.get("confirm_now") == "1"
        date        = request.form.get("date", "").strip()
        doc_number  = request.form.get("doc_number", "").strip()
        doc_type_id = request.form.get("doc_type_id") or None
        supplier    = request.form.get("supplier", "").strip()
        notes       = request.form.get("notes", "").strip()
        source_type = request.form.get("source_type", "income_doc")

        is_draft_save = save_as_draft and not confirm_now
        errors = []
        if not date and not is_draft_save:
            errors.append("Вкажіть дату")

        rows, row_errors = _parse_income_rows(request.form, is_draft_save, conn)
        errors.extend(row_errors)

        if not rows and not is_draft_save:
            errors.append("Додайте хоча б одну позицію")

        if errors:
            doc_items = conn.execute(
                "SELECT i.*, d.name AS item_name, d.unit_of_measure "
                "FROM income_doc_items i JOIN item_dictionary d ON d.id=i.item_id "
                "WHERE i.doc_id=?", (doc_id,)
            ).fetchall()
            conn.close()
            return render_template("warehouse/income_form.html",
                items=items, doc_types=doc_types, categories=CATEGORIES,
                norm_groups=norm_groups, errors=errors,
                form=request.form, doc=dict(doc), doc_items=[dict(r) for r in doc_items],
                editing_draft=True, doc_id=doc_id)

        scan_path     = doc["scan_path"]
        scan_orig     = doc["scan_original_name"]
        new_scan, new_orig = _save_scan_file(
            request.files.get("scan_file"), doc_number, date
        )
        if new_scan:
            scan_path, scan_orig = new_scan, new_orig

        # Оновлюємо заголовок і перезаписуємо позиції
        conn.execute("""
            UPDATE income_docs SET date=?, document_number=?, document_type_id=?,
                supplier=?, source_type=?, notes=?, scan_path=?, scan_original_name=?,
                updated_at=datetime('now','localtime')
            WHERE id=?
        """, (date, doc_number, doc_type_id, supplier, source_type,
              notes, scan_path, scan_orig, doc_id))
        conn.execute("DELETE FROM income_doc_items WHERE doc_id=?", (doc_id,))
        for row in rows:
            conn.execute("""
                INSERT INTO income_doc_items
                    (doc_id, item_id, quantity, price, category, nom_code, serial_numbers)
                VALUES (?,?,?,?,?,?,?)
            """, (doc_id, row["item_id"], row["quantity"], row["price"],
                  row["category"], row["nom_code"],
                  ",".join(row["serials"]) if row["serials"] else None))

        if confirm_now:
            # Переносимо в warehouse_income і видаляємо чернетку
            income_id = None
            for row in rows:
                cur = conn.execute("""
                    INSERT INTO warehouse_income
                        (date, document_number, document_type_id, supplier,
                         item_id, quantity, price, category, nom_code,
                         notes, source_type, status,
                         scan_path, scan_original_name, created_by, created_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,'confirmed',?,?,?,datetime('now','localtime'))
                """, (date, doc_number, doc_type_id, supplier,
                      row["item_id"], row["quantity"], row["price"],
                      row["category"], row["nom_code"], notes, source_type,
                      scan_path, scan_orig, session.get("user_id")))
                income_id = cur.lastrowid
                for sn in row["serials"]:
                    conn.execute("""
                        INSERT INTO item_serials (item_id, serial_number, warehouse_income_id, status)
                        VALUES (?,?,?,'stock')
                    """, (row["item_id"], sn, income_id))
            conn.execute("DELETE FROM income_doc_items WHERE doc_id=?", (doc_id,))
            conn.execute("DELETE FROM income_docs WHERE id=?", (doc_id,))
            conn.commit()
            log_action("add", "warehouse_income", income_id,
                       new_data={"date": date, "doc_number": doc_number, "rows": len(rows)})
            emit("warehouse.income", date=date, rows=rows)
            conn.close()
            from flask import flash
            flash("Прихід підтверджено та проведено на склад.", "success")
            return redirect(url_for("warehouse.income_list"))

        conn.commit()
        conn.close()
        from flask import flash
        flash("Чернетку оновлено.", "info")
        return redirect(url_for("warehouse.income_list"))

    # GET — завантажуємо поточні позиції
    doc_items = conn.execute(
        "SELECT i.*, d.name AS item_name, d.unit_of_measure, d.has_serial_number "
        "FROM income_doc_items i JOIN item_dictionary d ON d.id=i.item_id "
        "WHERE i.doc_id=?", (doc_id,)
    ).fetchall()
    conn.close()
    return render_template("warehouse/income_form.html",
        items=items, doc_types=doc_types, categories=CATEGORIES,
        norm_groups=norm_groups, errors=[],
        form=dict(doc), doc=dict(doc), doc_items=[dict(r) for r in doc_items],
        editing_draft=True, doc_id=doc_id)


@bp.route("/income/draft/<int:doc_id>/delete", methods=["POST"])
@login_required
def income_draft_delete(doc_id):
    conn = get_connection()
    conn.execute("DELETE FROM income_doc_items WHERE doc_id=?", (doc_id,))
    conn.execute("DELETE FROM income_docs WHERE id=?", (doc_id,))
    conn.commit()
    conn.close()
    from flask import flash
    flash("Чернетку видалено.", "info")
    return redirect(url_for("warehouse.income_list"))


def _get_income_doc_rows(conn, income_id):
    """Повертає (header_dict, [rows]) для підтвердженого приходу.
    Всі рядки документа визначаються по created_at першого запису."""
    head = conn.execute("""
        SELECT wi.*, dt.name AS doc_type_name
        FROM warehouse_income wi
        LEFT JOIN document_types dt ON dt.id = wi.document_type_id
        WHERE wi.id=?
    """, (income_id,)).fetchone()
    if not head:
        return None, []
    rows = conn.execute("""
        SELECT wi.*, d.name AS item_name, d.unit_of_measure, d.has_serial_number
        FROM warehouse_income wi
        JOIN item_dictionary d ON d.id = wi.item_id
        WHERE wi.created_at = ? AND COALESCE(wi.status,'confirmed')='confirmed'
        ORDER BY wi.id
    """, (head["created_at"],)).fetchall()
    return dict(head), [dict(r) for r in rows]


@bp.route("/income/<int:income_id>/view")
@login_required
def income_view(income_id):
    """Перегляд підтвердженого приходу."""
    conn = get_connection()
    head, rows = _get_income_doc_rows(conn, income_id)
    conn.close()
    if not head:
        from flask import flash
        flash("Прихід не знайдено", "danger")
        return redirect(url_for("warehouse.income_list"))
    return render_template("warehouse/income_view.html",
                           head=head, rows=rows, income_id=income_id)


@bp.route("/income/<int:income_id>/edit", methods=["GET", "POST"])
@login_required
def income_edit(income_id):
    """Редагування підтвердженого приходу."""
    conn = get_connection()
    head, existing_rows = _get_income_doc_rows(conn, income_id)
    if not head:
        conn.close()
        from flask import flash
        flash("Прихід не знайдено", "danger")
        return redirect(url_for("warehouse.income_list"))

    items       = _get_items_with_serial(conn)
    doc_types   = _get_doc_types(conn)
    norm_groups = _get_norm_groups(conn)

    if request.method == "POST":
        date        = request.form.get("date", "").strip()
        doc_number  = request.form.get("doc_number", "").strip()
        doc_type_id = request.form.get("doc_type_id") or None
        supplier    = request.form.get("supplier", "").strip()
        notes       = request.form.get("notes", "").strip()
        source_type = request.form.get("source_type", "income_doc")

        errors = []
        if not date:
            errors.append("Вкажіть дату")

        rows, row_errors = _parse_income_rows(request.form, False, conn)
        errors.extend(row_errors)

        if not rows:
            errors.append("Додайте хоча б одну позицію")

        if errors:
            conn.close()
            return render_template("warehouse/income_form.html",
                items=items, doc_types=doc_types, categories=CATEGORIES,
                norm_groups=norm_groups, errors=errors,
                form=request.form, doc=head,
                doc_items=existing_rows,
                editing_confirmed=True, income_id=income_id)

        scan_path  = head["scan_path"]
        scan_orig  = head["scan_original_name"]
        new_scan, new_orig = _save_scan_file(
            request.files.get("scan_file"), doc_number, date
        )
        if new_scan:
            scan_path, scan_orig = new_scan, new_orig

        # Визначаємо всі id рядків цього документа
        doc_ids = [r["id"] for r in existing_rows]

        # Видаляємо серійні номери прив'язані до цих рядків
        for did in doc_ids:
            conn.execute(
                "UPDATE item_serials SET status='deleted', warehouse_income_id=NULL "
                "WHERE warehouse_income_id=?", (did,)
            )
        # Видаляємо старі рядки
        for did in doc_ids:
            conn.execute("DELETE FROM warehouse_income WHERE id=?", (did,))

        # Вставляємо нові
        new_created_at = head["created_at"]  # зберігаємо оригінальний timestamp
        for row in rows:
            cur = conn.execute("""
                INSERT INTO warehouse_income
                    (date, document_number, document_type_id, supplier,
                     item_id, quantity, price, category, nom_code,
                     notes, source_type, status,
                     scan_path, scan_original_name, created_by, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,'confirmed',?,?,?,?)
            """, (date, doc_number, doc_type_id, supplier,
                  row["item_id"], row["quantity"], row["price"],
                  row["category"], row["nom_code"], notes, source_type,
                  scan_path, scan_orig, session.get("user_id"), new_created_at))
            new_id = cur.lastrowid
            for sn in row["serials"]:
                conn.execute("""
                    INSERT INTO item_serials (item_id, serial_number, warehouse_income_id, status)
                    VALUES (?,?,?,'stock')
                """, (row["item_id"], sn, new_id))

        conn.commit()
        log_action("edit", "warehouse_income", income_id,
                   new_data={"date": date, "doc_number": doc_number, "rows": len(rows)})
        conn.close()
        from flask import flash
        flash("Прихід оновлено.", "success")
        return redirect(url_for("warehouse.income_list"))

    conn.close()
    return render_template("warehouse/income_form.html",
        items=items, doc_types=doc_types, categories=CATEGORIES,
        norm_groups=norm_groups, errors=[],
        form=head, doc=head,
        doc_items=existing_rows,
        editing_confirmed=True, income_id=income_id)


# ── API: приходи конкретного майна ──────────────────────────────
@bp.route("/api/item/<int:item_id>/incomes")
@login_required
def api_item_incomes(item_id):
    conn = get_connection()
    item = conn.execute(
        "SELECT id, name, unit_of_measure FROM item_dictionary WHERE id=?", (item_id,)
    ).fetchone()
    if not item:
        conn.close()
        return jsonify({"error": "Не знайдено"}), 404
    rows = conn.execute("""
        SELECT wi.id, wi.date, wi.document_number, wi.supplier,
               dt.name AS doc_type_name, wi.quantity, wi.price, wi.category,
               wi.notes, wi.created_at
        FROM warehouse_income wi
        LEFT JOIN document_types dt ON dt.id = wi.document_type_id
        WHERE wi.item_id = ? AND COALESCE(wi.status,'confirmed')='confirmed'
        ORDER BY wi.date DESC, wi.id DESC
    """, (item_id,)).fetchall()
    conn.close()
    return jsonify({
        "item": {"id": item["id"], "name": item["name"], "unit": item["unit_of_measure"]},
        "rows": [dict(r) for r in rows],
    })


@bp.route("/income/<int:income_id>/delete", methods=["POST"])
@login_required
def income_delete(income_id):
    conn = get_connection()
    row = conn.execute(
        "SELECT * FROM warehouse_income WHERE id = ?", (income_id,)
    ).fetchone()
    if row:
        conn.execute(
            "UPDATE item_serials SET status='deleted' WHERE warehouse_income_id = ?",
            (income_id,)
        )
        conn.execute("DELETE FROM warehouse_income WHERE id = ?", (income_id,))
        conn.commit()
        log_action("delete", "warehouse_income", income_id, old_data=dict(row))
    conn.close()
    return redirect(url_for("warehouse.income_list"))


@bp.route("/income/<int:income_id>/scan")
@login_required
def income_scan(income_id):
    """Відкрити/завантажити скан документа приходу."""
    from flask import send_file, abort
    conn = get_connection()
    row = conn.execute(
        "SELECT scan_path, scan_original_name FROM warehouse_income WHERE id=?",
        (income_id,)
    ).fetchone()
    conn.close()
    if not row or not row["scan_path"]:
        abort(404)
    scan_path = row["scan_path"]
    if not os.path.exists(scan_path):
        abort(404)
    return send_file(
        scan_path,
        as_attachment=False,
        download_name=row["scan_original_name"] or os.path.basename(scan_path),
    )


# ─────────────────────────────────────────────────────────────
#  ІНВЕНТАРИЗАЦІЯ — список
# ─────────────────────────────────────────────────────────────

@bp.route("/inventory/")
@login_required
def inventory_list():
    conn = get_connection()
    rows = conn.execute("""
        SELECT inv.*,
               u.full_name AS created_by_name
        FROM inventories inv
        LEFT JOIN users u ON inv.created_by = u.id
        ORDER BY inv.date DESC, inv.id DESC
    """).fetchall()
    conn.close()
    return render_template("warehouse/inventory_list.html", rows=rows)


# ─────────────────────────────────────────────────────────────
#  ІНВЕНТАРИЗАЦІЯ — нова
# ─────────────────────────────────────────────────────────────

@bp.route("/inventory/new", methods=["GET", "POST"])
@login_required
def inventory_new():
    conn = get_connection()

    if request.method == "POST":
        date     = request.form.get("date", "").strip()
        notes    = request.form.get("notes", "").strip()
        members  = request.form.get("members", "").strip()

        if not date:
            stock = _get_stock(conn)
            conn.close()
            return render_template(
                "warehouse/inventory_form.html",
                stock=stock, categories=CATEGORIES,
                errors=["Вкажіть дату інвентаризації"],
                form=request.form,
            )

        # Зберігаємо шапку
        cur = conn.execute("""
            INSERT INTO inventories (date, notes, commission_members,
                                     status, created_by, created_at)
            VALUES (?, ?, ?, 'draft', ?, datetime('now','localtime'))
        """, (date, notes, members, session.get("user_id")))
        inv_id = cur.lastrowid

        # Рядки — з форми (фактична кількість)
        item_ids    = request.form.getlist("item_id[]")
        cats        = request.form.getlist("category[]")
        prices      = request.form.getlist("price[]")
        qty_expect  = request.form.getlist("qty_expected[]")
        qty_fact    = request.form.getlist("qty_actual[]")
        item_names  = request.form.getlist("item_name[]")
        units       = request.form.getlist("unit_of_measure[]")

        for i, item_id in enumerate(item_ids):
            try:
                qe = float(qty_expect[i]) if qty_expect[i] else 0
                qf = float(qty_fact[i])   if qty_fact[i]   else 0
                pr = float(prices[i])     if prices[i]     else 0
                cat  = cats[i]       if i < len(cats)       else "I"
                name = item_names[i] if i < len(item_names) else ""
                unit = units[i]      if i < len(units)      else "шт"
            except (ValueError, IndexError):
                continue

            conn.execute("""
                INSERT INTO inventory_items
                    (inventory_id, item_id, item_name_snapshot,
                     unit_of_measure, category, price,
                     qty_expected, qty_actual)
                VALUES (?,?,?,?,?,?,?,?)
            """, (inv_id, item_id, name, unit, cat, pr, qe, qf))

        conn.execute(
            "UPDATE inventories SET status='done' WHERE id = ?", (inv_id,)
        )
        conn.commit()
        log_action("add", "inventories", inv_id,
                   new_data={"date": date, "items": len(item_ids)})
        conn.close()
        return redirect(url_for("warehouse.inventory_view", inv_id=inv_id))

    # GET — завантажити поточні залишки як основу
    stock = _get_stock(conn)
    conn.close()
    today = datetime.now().strftime("%Y-%m-%d")
    return render_template(
        "warehouse/inventory_form.html",
        stock=stock, categories=CATEGORIES,
        errors=[],
        form={"date": today},
    )


# ─────────────────────────────────────────────────────────────
#  ІНВЕНТАРИЗАЦІЯ — перегляд
# ─────────────────────────────────────────────────────────────

@bp.route("/inventory/<int:inv_id>")
@login_required
def inventory_view(inv_id):
    conn = get_connection()
    inv = conn.execute(
        "SELECT * FROM inventories WHERE id = ?", (inv_id,)
    ).fetchone()
    if not inv:
        conn.close()
        return redirect(url_for("warehouse.inventory_list"))

    items = conn.execute(
        "SELECT * FROM inventory_items WHERE inventory_id = ? ORDER BY item_name_snapshot, category",
        (inv_id,)
    ).fetchall()
    conn.close()

    total_expected = sum(r["qty_expected"] * r["price"] for r in items)
    total_actual   = sum(r["qty_actual"]   * r["price"] for r in items)

    return render_template(
        "warehouse/inventory_view.html",
        inv=inv, items=items,
        total_expected=total_expected,
        total_actual=total_actual,
    )


@bp.route("/inventory/<int:inv_id>/delete", methods=["POST"])
@login_required
def inventory_delete(inv_id):
    conn = get_connection()
    conn.execute("DELETE FROM inventory_items WHERE inventory_id = ?", (inv_id,))
    conn.execute("DELETE FROM inventories WHERE id = ?", (inv_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("warehouse.inventory_list"))


# ─────────────────────────────────────────────────────────────
#  ДРУК — HTML для print / Excel / PDF
# ─────────────────────────────────────────────────────────────

@bp.route("/inventory/<int:inv_id>/print")
@login_required
def inventory_print(inv_id):
    conn = get_connection()
    inv   = conn.execute("SELECT * FROM inventories WHERE id = ?", (inv_id,)).fetchone()
    items = conn.execute(
        "SELECT * FROM inventory_items WHERE inventory_id = ? ORDER BY item_name_snapshot, category",
        (inv_id,)
    ).fetchall()
    from core.settings import get_setting
    unit_name = get_setting("company_name", "") or ""
    conn.close()

    total_expected = sum(r["qty_expected"] * r["price"] for r in items)
    total_actual   = sum(r["qty_actual"]   * r["price"] for r in items)

    return render_template(
        "warehouse/inventory_print.html",
        inv=inv, items=items,
        unit_name=unit_name,
        total_expected=total_expected,
        total_actual=total_actual,
    )


@bp.route("/inventory/<int:inv_id>/export/xlsx")
@login_required
def inventory_export_xlsx(inv_id):
    """Вивантаження інвентаризації в Excel."""
    conn = get_connection()
    inv   = conn.execute("SELECT * FROM inventories WHERE id = ?", (inv_id,)).fetchone()
    items = conn.execute(
        "SELECT * FROM inventory_items WHERE inventory_id = ? ORDER BY item_name_snapshot, category",
        (inv_id,)
    ).fetchall()
    from core.settings import get_setting
    unit_name = get_setting("company_name", "") or ""
    conn.close()

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        return "Бібліотека openpyxl не встановлена. Встановіть: pip install openpyxl", 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Інвентаризація"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Заголовок
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Відомість наявності матеріальних цінностей на складі"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = f"{unit_name}   Дата: {inv['date']}"
    ws["A2"].alignment = center

    if inv["commission_members"]:
        ws.merge_cells("A3:H3")
        ws["A3"] = f"Комісія: {inv['commission_members']}"
        ws["A3"].alignment = Alignment(wrap_text=True)

    # Шапка таблиці
    headers = ["№", "Найменування", "Кат.", "Ціна, грн",
               "По обліку", "Фактично", "Різниця", "Сума факт., грн"]
    widths  = [5, 40, 8, 12, 12, 12, 12, 16]
    header_row = 5
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = bold
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[header_row].height = 30

    # Рядки даних
    for i, row in enumerate(items, 1):
        diff    = round(row["qty_actual"] - row["qty_expected"], 4)
        sum_val = round(row["qty_actual"] * row["price"], 2)
        data = [
            i,
            row["item_name_snapshot"],
            row["category"],
            row["price"],
            row["qty_expected"],
            row["qty_actual"],
            diff if diff != 0 else "",
            sum_val,
        ]
        r = header_row + i
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col != 2 else "left",
                vertical="center", wrap_text=True
            )

    # Підсумок
    total_row = header_row + len(items) + 1
    total_actual = sum(r["qty_actual"] * r["price"] for r in items)
    ws.cell(row=total_row, column=1, value="Разом:").font = bold
    ws.merge_cells(f"A{total_row}:G{total_row}")
    ws["A" + str(total_row)].alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=8, value=round(total_actual, 2)).font = bold

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    date_str = (inv["date"] or "").replace("-", "")
    filename = f"inventory_{date_str}.xlsx"

    resp = make_response(buf.read())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp


@bp.route("/inventory/<int:inv_id>/export/pdf")
@login_required
def inventory_export_pdf(inv_id):
    """PDF через print-шаблон (відкриває сторінку для друку браузером)."""
    return redirect(url_for("warehouse.inventory_print", inv_id=inv_id))


# ─────────────────────────────────────────────────────────────
#  JSON API — для динамічних форм
# ─────────────────────────────────────────────────────────────

@bp.route("/api/item/<int:item_id>")
@login_required
def api_item(item_id):
    """Повертає дані позиції словника + доступні ціни/категорії зі складу."""
    conn = get_connection()
    row = conn.execute(
        "SELECT id, name, unit_of_measure, has_serial_number, is_inventory "
        "FROM item_dictionary WHERE id = ?",
        (item_id,)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "not found"}), 404

    # Доступні партії зі складу (qty_free > 0)
    stock = _get_stock(conn, only_positive=True)
    conn.close()

    prices = [
        {"price": s["price"], "category": s["category"], "qty_free": s["qty_free"]}
        for s in stock if s["item_id"] == item_id
    ]

    result = dict(row)
    result["stock_prices"] = prices
    # Перша доступна ціна для автопідстановки
    if prices:
        result["default_price"] = prices[0]["price"]
        result["default_category"] = prices[0]["category"]
    else:
        result["default_price"] = 0
        result["default_category"] = "I"
    return jsonify(result)
