"""
modules/import_cards/parser.py — парсинг Excel-карток А5027 → JSON

Структура Excel (1 аркуш = 1 особа):
  рядок 1,  кол A   → "КАРТКА № 173"
  рядок 5,  кол A–L → розміри
  рядок 9,  кол D   → звання
  рядок 9,  кол E   → ПІБ
  рядок 9,  кол K   → підрозділ
  рядок 11, кол F–V → назви РВ/накладних
  рядки 12–71:
      кол B   → найменування предмета
      кол E   → qty+date: "1  03.2022"
      кол F–V → кількість по кожному РВ
      остання числова кол. → ціна

Аркуші "Зміст" і "Зразок" — пропускаються.
Блок атестату (кол X+) — ігнорується.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import column_index_from_string


# ── Константи ─────────────────────────────────────────────────────────────────

SKIP_SHEETS = {"зміст", "зразок"}   # нижній регістр для порівняння

ROW_CARD_NUMBER = 1
ROW_SIZES       = 5
ROW_PERSON      = 9
ROW_ENROLL      = 10   # рядок з атестатом (нам не потрібен, але можна логувати)
ROW_DOCS        = 11   # заголовки РВ/накладних
ROW_ITEMS_START = 12
ROW_ITEMS_END   = 71

COL_CARD_NUMBER = 1    # A
COL_RANK        = 4    # D
COL_NAME        = 5    # E
COL_UNIT        = 11   # K
COL_ITEM_NUM    = 1    # A
COL_ITEM_NAME   = 2    # B
COL_ITEM_ATTEST = 5    # E  (qty + date з атестату)
COL_DOCS_START  = 6    # F  (перший РВ/накладна)
COL_DOCS_END    = 22   # V  (максимум 17 колонок: F..V)

# Колонки розмірів: (мітка, номер колонки)
SIZE_COLS = [
    ("size_head",       1),   # A — Голова
    ("size_height",     3),   # C — Зріст
    ("size_underwear",  5),   # E — Білизна
    ("size_suit",       7),   # G — Костюм
    ("size_jacket",     9),   # I — Куртка
    ("size_pants",      11),  # K — Штани
    ("size_shoes",      13),  # M — Взуття
]

# Regex для парсингу поля атестату: "1  03.2022" або "2 07.2025"
_RE_ATTEST = re.compile(r"^\s*(\d+(?:[.,]\d+)?)\s+(\d{2})\.(\d{4})\s*$")

# Regex для парсингу назви РВ/накладної.
# Підтримує формати:
#   "РВ №35 від 10.12.2025"
#   "РВ №150  18.03.2025"   (пробіли без "від")
#   "накл. № 14  02.06.2025"
#   "накл. №253  27.05.2025"
#   "НЛ №12 від 05.03.2025"
_RE_DOC = re.compile(
    r"(?:РВ|НЛ|накл\.?)\s*[№#]?\s*(\d+)\s+(?:від\s+)?(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})",
    re.IGNORECASE,
)


# ── Допоміжні ─────────────────────────────────────────────────────────────────

def _cell_str(ws, row: int, col: int) -> str:
    """Повертає рядкове значення клітинки (merged або звичайної), trimmed."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()


def _cell_float(ws, row: int, col: int) -> Optional[float]:
    """Повертає числове значення або None."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def _parse_card_number(text: str) -> str:
    """'КАРТКА № 173' → '173'. Якщо не вдалось — повертає ''."""
    m = re.search(r"№\s*(\S+)", text)
    return m.group(1) if m else ""


def _parse_name(text: str) -> tuple[str, str, str]:
    """
    'ІВАНОВ Ярослав Васильович' → (last, first, middle).
    Повертає ('', '', '') якщо не вдалось розпарсити.
    """
    parts = text.split()
    if not parts:
        return "", "", ""
    last   = parts[0].capitalize() if parts[0].isupper() else parts[0]
    first  = parts[1] if len(parts) > 1 else ""
    middle = parts[2] if len(parts) > 2 else ""
    # Прізвище зберігаємо як є (КАПС або мікс — нормалізуємо до Title)
    last = last.title()
    return last, first, middle


def _parse_attest_cell(text: str) -> tuple[Optional[float], Optional[str]]:
    """
    '1  03.2022' → (1.0, '2022-03-01')
    '2 07.2025'  → (2.0, '2025-07-01')
    Повертає (None, None) якщо не вдалось.
    """
    m = _RE_ATTEST.match(text)
    if not m:
        return None, None
    qty   = float(m.group(1).replace(",", "."))
    month = int(m.group(2))
    year  = int(m.group(3))
    date  = f"{year:04d}-{month:02d}-01"
    return qty, date


def _parse_doc_header(text: str) -> Optional[dict]:
    """
    'РВ №35 від 10.12.2025' → {'raw': '...', 'number': '35', 'date': '2025-12-10', 'doc_type': 'rv'}
    'НЛ №12 від 05.03.2025' → {'...', 'doc_type': 'invoice'}
    None якщо не вдалось.
    """
    if not text:
        return None
    m = _RE_DOC.search(text)
    if not m:
        return None

    number  = m.group(1).strip(".")
    date_raw = m.group(2)

    # Нормалізуємо дату до YYYY-MM-DD
    date_iso = _normalize_date(date_raw)

    doc_type = "rv" if re.search(r"\bРВ\b", text, re.IGNORECASE) else "invoice"

    return {
        "raw":      text.strip(),
        "number":   number,
        "date":     date_iso,
        "doc_type": doc_type,
    }


def _normalize_date(date_str: str) -> Optional[str]:
    """
    '10.12.2025' або '10/12/2025' або '10-12-2025' → '2025-12-10'
    Повертає None якщо не вдалось.
    """
    date_str = date_str.strip()
    m = re.match(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", date_str)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    try:
        return f"{y:04d}-{mo:02d}-{d:02d}"
    except ValueError:
        return None


def _find_price_col(ws, row: int, doc_count: int) -> Optional[float]:
    """
    Ціна — перша числова колонка після колонок РВ.
    doc_count — скільки РВ знайдено в рядку 11.
    """
    price_col_start = COL_DOCS_START + doc_count
    for c in range(price_col_start, price_col_start + 6):
        val = _cell_float(ws, row, c)
        if val and val > 0:
            return val
    return None


# ── Парсинг одного аркуша ──────────────────────────────────────────────────────

def _parse_sheet(ws) -> dict:
    """
    Парсить один аркуш → dict з даними особи.

    Структура повернення:
    {
        "card_number": "173",
        "last_name": "Іванов",
        "first_name": "Ярослав",
        "middle_name": "Васильович",
        "rank": "старший солдат",
        "unit_raw": "А5215",
        "sizes": {"size_head": "57", "size_height": "3", ...},
        "docs": [
            {"raw": "РВ №35 від 10.12.2025", "number": "35",
             "date": "2025-12-10", "doc_type": "rv"},
            ...
        ],
        "items": [
            {
                "row_num": 1,
                "name_raw": "Кепі бойове (кашкет ЛП)",
                "attest_qty": 1.0,
                "attest_date": "2022-03-01",
                "price": 699.9,
                "doc_quantities": [1, None, None, ...]  # індекс = індекс docs[]
            },
            ...
        ],
        "parse_warnings": ["..."]   # непорожній якщо щось підозріле
    }
    """
    warnings = []

    # ── Номер картки ──
    card_raw = _cell_str(ws, ROW_CARD_NUMBER, COL_CARD_NUMBER)
    card_number = _parse_card_number(card_raw)
    if not card_number:
        warnings.append(f"Не вдалось зчитати номер картки: '{card_raw}'")

    # ── ПІБ і звання ──
    rank_raw = _cell_str(ws, ROW_PERSON, COL_RANK)
    name_raw = _cell_str(ws, ROW_PERSON, COL_NAME)
    unit_raw = _cell_str(ws, ROW_PERSON, COL_UNIT)

    # Іноді ПІБ і звання — в одній клітинці, спробуємо обидва варіанти
    if not name_raw and rank_raw:
        # Можливо ПІБ у полі звання через merged cells
        name_raw = rank_raw
        rank_raw = ""

    last_name, first_name, middle_name = _parse_name(name_raw)
    if not last_name:
        warnings.append(f"Не вдалось зчитати ПІБ: '{name_raw}'")

    # ── Розміри (рядок 5) ──
    sizes = {}
    for field, col in SIZE_COLS:
        val = _cell_str(ws, ROW_SIZES, col)
        if val:
            sizes[field] = val

    # ── Заголовки документів (рядок 11, колонки F–V) ──
    # Пропускаємо порожні заготовки: "РВ №", "накл. № ДАТА" без конкретного номера/дати
    _TEMPLATE_STUBS = re.compile(
        r"^(рв\s*[№#]?|накл\.?\s*[№#]?\s*(дата)?|нл\s*[№#]?)$",
        re.IGNORECASE,
    )
    docs = []
    for col in range(COL_DOCS_START, COL_DOCS_END + 1):
        text = _cell_str(ws, ROW_DOCS, col)
        if not text:
            continue
        if _TEMPLATE_STUBS.match(text.strip()):
            continue  # порожня заготовка — пропускаємо без попередження
        doc = _parse_doc_header(text)
        if doc:
            doc["col_index"] = col  # зберігаємо для матчингу з рядками майна
            docs.append(doc)
        else:
            warnings.append(f"Невідомий формат документа в рядку 11, кол {col}: '{text}'")

    doc_count = len(docs)

    # ── Рядки майна (12–71) ──
    items = []
    for row in range(ROW_ITEMS_START, ROW_ITEMS_END + 1):
        # Назва майна
        name = _cell_str(ws, row, COL_ITEM_NAME)
        if not name:
            continue  # порожній рядок — пропускаємо

        # Поле атестату (qty + date)
        attest_text = _cell_str(ws, row, COL_ITEM_ATTEST)
        attest_qty, attest_date = _parse_attest_cell(attest_text)
        if attest_text and attest_qty is None:
            # Може бути просто число без дати
            try:
                attest_qty = float(attest_text.replace(",", "."))
                attest_date = None
            except ValueError:
                warnings.append(f"Рядок {row}: не розпарсили атестат '{attest_text}'")

        # Кількості по РВ (індекс відповідає індексу в docs[])
        doc_quantities: list[Optional[float]] = []
        for doc in docs:
            qty = _cell_float(ws, row, doc["col_index"])
            doc_quantities.append(qty)

        # Ціна
        price = _find_price_col(ws, row, doc_count)

        # Пропускаємо якщо взагалі немає даних про видачу
        has_any = (attest_qty is not None) or any(q for q in doc_quantities if q)
        if not has_any:
            continue

        row_num_val = _cell_str(ws, row, COL_ITEM_NUM)
        items.append({
            "row_num":       row_num_val,
            "name_raw":      name,
            "attest_qty":    attest_qty,
            "attest_date":   attest_date,
            "price":         price,
            "doc_quantities": doc_quantities,
        })

    return {
        "card_number":  card_number,
        "last_name":    last_name,
        "first_name":   first_name,
        "middle_name":  middle_name,
        "rank":         rank_raw,
        "unit_raw":     unit_raw,
        "sizes":        sizes,
        "docs":         docs,
        "items":        items,
        "parse_warnings": warnings,
    }


# ── Головна функція ────────────────────────────────────────────────────────────

def parse_file(filepath: str | Path) -> list[dict]:
    """
    Парсить Excel файл з картками А5027.

    Повертає список dict — по одному на кожну знайдену картку.
    Аркуші 'Зміст' і 'Зразок' пропускаються автоматично.

    Raises: ValueError якщо файл не читається.
    """
    path = Path(filepath)
    if not path.exists():
        raise ValueError(f"Файл не знайдено: {path}")

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Не вдалось відкрити файл: {e}") from e

    results = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        try:
            card = _parse_sheet(ws)
            card["sheet_name"] = sheet_name
            results.append(card)
        except Exception as e:
            results.append({
                "sheet_name":     sheet_name,
                "card_number":    "",
                "last_name":      sheet_name,
                "first_name":     "",
                "middle_name":    "",
                "rank":           "",
                "unit_raw":       "",
                "sizes":          {},
                "docs":           [],
                "items":          [],
                "parse_warnings": [f"Критична помилка парсингу: {e}"],
            })

    wb.close()
    return results
